from fastapi import FastAPI, Depends, UploadFile, File, HTTPException, status
from fastapi.responses import HTMLResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
from sqlalchemy import select, func
from openpyxl import load_workbook
import csv
import io

from .database import get_db
from . import models, schemas
from .tasks import build_messages_for_campaign, send_campaign_job
from .security import hash_password, verify_password, create_access_token, decode_token

app = FastAPI(
    title="Система массовых email-рассылок S8 Airlines",
    description="""
API для дипломного проекта по управлению массовыми email-рассылками.

Возможности системы:
- регистрация и вход пользователей;
- импорт контактов из CSV;
- импорт контактов из Excel (.xlsx);
- создание шаблонов писем;
- создание и запуск кампаний;
- формирование сообщений;
- получение отчётов по отправке;
- отписка получателей от рассылки.
""",
    version="1.0.0",
)

auth_scheme = HTTPBearer()


def get_current_user(
    creds: HTTPAuthorizationCredentials = Depends(auth_scheme),
    db: Session = Depends(get_db),
) -> models.User:
    """
    Получение текущего пользователя по JWT-токену.
    """
    try:
        payload = decode_token(creds.credentials)
        user_id = int(payload["sub"])
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Некорректный токен авторизации",
        )

    u = db.get(models.User, user_id)
    if not u or not u.is_active:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Пользователь не найден или неактивен",
        )
    return u


# ---------- ГЛАВНАЯ СТРАНИЦА ----------
@app.get("/", response_class=HTMLResponse, summary="Главная страница", tags=["Главная"])
def root():
    return """
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8" />
        <meta name="viewport" content="width=device-width, initial-scale=1.0" />
        <title>Система массовых рассылок S8 Airlines</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                background: #f4f7fb;
                margin: 0;
                padding: 0;
                color: #222;
            }
            .container {
                max-width: 1100px;
                margin: 40px auto;
                background: white;
                border-radius: 16px;
                padding: 32px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.08);
            }
            h1 {
                margin-top: 0;
                color: #0b5cab;
            }
            p {
                line-height: 1.6;
            }
            .grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
                gap: 16px;
                margin-top: 24px;
            }
            .card {
                background: #f8fbff;
                border: 1px solid #d8e7f7;
                border-radius: 12px;
                padding: 18px;
            }
            .card h3 {
                margin-top: 0;
                color: #124b87;
            }
            .btn {
                display: inline-block;
                margin-top: 10px;
                padding: 10px 14px;
                background: #0b5cab;
                color: white;
                text-decoration: none;
                border-radius: 8px;
                font-weight: bold;
            }
            .btn:hover {
                background: #08498a;
            }
            .note {
                margin-top: 28px;
                padding: 16px;
                background: #fff8e8;
                border: 1px solid #f0d999;
                border-radius: 10px;
            }
            code {
                background: #eef3f8;
                padding: 2px 6px;
                border-radius: 6px;
            }
            ul {
                padding-left: 20px;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Система массовых email-рассылок S8 Airlines</h1>
            <p>
                Сервер успешно работает. Это стартовая страница backend-приложения.
                Ниже находятся быстрые переходы к основным разделам системы.
            </p>

            <div class="grid">
                <div class="card">
                    <h3>Документация API</h3>
                    <p>Открыть Swagger-интерфейс для тестирования всех методов системы.</p>
                    <a class="btn" href="/docs">Открыть /docs</a>
                </div>

                <div class="card">
                    <h3>OpenAPI схема</h3>
                    <p>Техническое JSON-описание всех маршрутов API.</p>
                    <a class="btn" href="/openapi.json">Открыть openapi.json</a>
                </div>

                <div class="card">
                    <h3>Регистрация пользователя</h3>
                    <p>Создание нового пользователя через Swagger.</p>
                    <a class="btn" href="/docs#/Авторизация/register_auth_register_post">Открыть</a>
                </div>

                <div class="card">
                    <h3>Вход в систему</h3>
                    <p>Получение JWT-токена для работы с защищёнными методами.</p>
                    <a class="btn" href="/docs#/Авторизация/login_auth_login_post">Открыть</a>
                </div>

                <div class="card">
                    <h3>Импорт контактов из CSV</h3>
                    <p>Загрузка CSV-файла с контактами для рассылки.</p>
                    <a class="btn" href="/docs#/Контакты/import_contacts_csv_contacts_import_csv_post">Открыть</a>
                </div>

                <div class="card">
                    <h3>Импорт контактов из Excel</h3>
                    <p>Загрузка Excel-файла .xlsx со списком получателей.</p>
                    <a class="btn" href="/docs#/Контакты/import_contacts_excel_contacts_import_excel_post">Открыть</a>
                </div>

                <div class="card">
                    <h3>Шаблоны писем</h3>
                    <p>Создание и просмотр шаблонов рассылки.</p>
                    <a class="btn" href="/docs#/Шаблоны">Открыть</a>
                </div>

                <div class="card">
                    <h3>Кампании</h3>
                    <p>Создание, подготовка, отправка и просмотр кампаний.</p>
                    <a class="btn" href="/docs#/Кампании">Открыть</a>
                </div>

                <div class="card">
                    <h3>Отписка</h3>
                    <p>Ссылка для отписки получателя от email-рассылки.</p>
                    <a class="btn" href="/docs#/Отписка">Открыть</a>
                </div>
            </div>

            <div class="note">
                <strong>Важно:</strong>
                <ul>
                    <li><code>GET</code>-маршруты можно открывать напрямую в браузере.</li>
                    <li><code>POST</code>-маршруты, такие как регистрация, вход, создание шаблона и кампании, нужно вызывать через <code>/docs</code> или через графическое приложение.</li>
                    <li>Для защищённых маршрутов сначала нужно получить токен через вход и нажать кнопку <strong>Authorize</strong> в Swagger.</li>
                </ul>
            </div>
        </div>
    </body>
    </html>
    """


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    raise HTTPException(status_code=204)


# ---------- АВТОРИЗАЦИЯ ----------
@app.post("/auth/register", summary="Регистрация пользователя", tags=["Авторизация"])
def register(payload: schemas.UserRegister, db: Session = Depends(get_db)):
    """
    Регистрация нового пользователя.
    """
    if len(payload.password) < 6:
        raise HTTPException(
            status_code=400,
            detail="Пароль слишком короткий. Минимум 6 символов.",
        )

    exists = db.scalar(select(models.User).where(models.User.email == payload.email))
    if exists:
        raise HTTPException(status_code=400, detail="Пользователь уже существует")

    u = models.User(
        email=payload.email,
        password_hash=hash_password(payload.password),
        role=payload.role,
    )
    db.add(u)
    db.commit()
    db.refresh(u)

    return {
        "id": u.id,
        "email": u.email,
        "role": u.role,
        "message": "Пользователь успешно зарегистрирован",
    }


@app.post("/auth/login", response_model=schemas.TokenOut, summary="Вход в систему", tags=["Авторизация"])
def login(payload: schemas.UserLogin, db: Session = Depends(get_db)):
    """
    Вход пользователя в систему и выдача JWT-токена.
    """
    u = db.scalar(select(models.User).where(models.User.email == payload.email))
    if not u or not verify_password(payload.password, u.password_hash):
        raise HTTPException(
            status_code=401,
            detail="Неверный email или пароль",
        )

    token = create_access_token(user_id=u.id, role=u.role)
    return schemas.TokenOut(access_token=token)


# ---------- КОНТАКТЫ ----------
@app.post("/contacts/import_csv", summary="Импорт контактов из CSV", tags=["Контакты"])
def import_contacts_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Импорт контактов из CSV-файла.
    """
    try:
        content = file.file.read().decode("utf-8-sig")
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать CSV-файл")

    reader = csv.DictReader(io.StringIO(content))
    created = 0
    skipped = 0

    for row in reader:
        email = (row.get("email") or "").strip()
        if not email:
            skipped += 1
            continue

        full_name = (row.get("full_name") or "").strip() or None
        ext_id = (row.get("external_client_id") or "").strip() or None
        consent_raw = (row.get("consent") or "true").strip().lower()
        consent = consent_raw in ("1", "true", "yes", "y", "да")

        c = models.Contact(
            email=email,
            full_name=full_name,
            external_client_id=ext_id,
            consent=consent,
        )
        db.add(c)

        try:
            db.commit()
            created += 1
        except Exception:
            db.rollback()
            skipped += 1

    return {
        "message": "Импорт контактов из CSV завершён",
        "created": created,
        "skipped": skipped,
    }


@app.post("/contacts/import_excel", summary="Импорт контактов из Excel", tags=["Контакты"])
def import_contacts_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Импорт контактов из Excel-файла (.xlsx).
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Поддерживаются только файлы формата .xlsx")

    try:
        workbook = load_workbook(file.file)
        sheet = workbook.active
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel-файл пустой")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers)}

    if "email" not in header_map:
        raise HTTPException(status_code=400, detail="В Excel-файле должна быть колонка 'email'")

    created = 0
    skipped = 0

    for row in rows[1:]:
        email_idx = header_map["email"]
        email = ""
        if email_idx < len(row) and row[email_idx] is not None:
            email = str(row[email_idx]).strip()

        if not email:
            skipped += 1
            continue

        full_name = None
        if "full_name" in header_map and header_map["full_name"] < len(row):
            value = row[header_map["full_name"]]
            full_name = str(value).strip() if value else None

        external_client_id = None
        if "external_client_id" in header_map and header_map["external_client_id"] < len(row):
            value = row[header_map["external_client_id"]]
            external_client_id = str(value).strip() if value else None

        consent = True
        if "consent" in header_map and header_map["consent"] < len(row):
            value = row[header_map["consent"]]
            consent_raw = str(value).strip().lower() if value is not None else "true"
            consent = consent_raw in ("1", "true", "yes", "y", "да")

        c = models.Contact(
            email=email,
            full_name=full_name,
            external_client_id=external_client_id,
            consent=consent,
        )
        db.add(c)

        try:
            db.commit()
            created += 1
        except Exception:
            db.rollback()
            skipped += 1

    return {
        "message": "Импорт контактов из Excel завершён",
        "created": created,
        "skipped": skipped,
    }


# ---------- ШАБЛОНЫ ----------
@app.post("/templates", response_model=int, summary="Создать шаблон письма", tags=["Шаблоны"])
def create_template(
    payload: schemas.TemplateCreate,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Создание нового шаблона письма.
    """
    t = models.Template(**payload.model_dump())
    db.add(t)
    db.commit()
    db.refresh(t)
    return t.id


@app.get("/templates", summary="Получить список шаблонов", tags=["Шаблоны"])
def list_templates(
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Возвращает список доступных шаблонов.
    """
    rows = db.execute(
        select(models.Template.id, models.Template.name).order_by(models.Template.id.desc())
    ).all()
    return [{"id": r.id, "name": r.name} for r in rows]


# ---------- КАМПАНИИ ----------
@app.post("/campaigns", response_model=int, summary="Создать кампанию", tags=["Кампании"])
def create_campaign(
    payload: schemas.CampaignCreate,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Создание новой кампании по выбранному шаблону.
    """
    tpl = db.get(models.Template, payload.template_id)
    if not tpl:
        raise HTTPException(status_code=400, detail="Шаблон не найден")

    camp = models.Campaign(
        name=payload.name,
        description=payload.description,
        template_id=payload.template_id,
        created_by=user.id,
    )
    db.add(camp)
    db.commit()
    db.refresh(camp)
    return camp.id


@app.get("/campaigns", summary="Получить список кампаний", tags=["Кампании"])
def list_campaigns(
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Возвращает список рассылочных кампаний.
    """
    rows = db.execute(
        select(models.Campaign.id, models.Campaign.name, models.Campaign.status).order_by(models.Campaign.id.desc())
    ).all()
    return [{"id": r.id, "name": r.name, "status": r.status} for r in rows]


@app.post("/campaigns/{campaign_id}/prepare", summary="Подготовить кампанию", tags=["Кампании"])
def prepare_campaign(
    campaign_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Формирование сообщений для выбранной кампании.
    """
    camp = db.get(models.Campaign, campaign_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Кампания не найдена")

    created = build_messages_for_campaign(db, campaign_id)
    camp.status = "queued"
    db.commit()

    return {
        "message": "Кампания подготовлена",
        "campaign_id": campaign_id,
        "messages_created": created,
    }


@app.post("/campaigns/{campaign_id}/send", summary="Отправить кампанию", tags=["Кампании"])
def send_campaign_direct(
    campaign_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Прямая отправка кампании без очереди и без отдельного воркера.
    """
    camp = db.get(models.Campaign, campaign_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Кампания не найдена")

    if camp.status not in ("queued", "draft"):
        raise HTTPException(
            status_code=400,
            detail=f"Статус кампании: {camp.status}. Ожидается queued или draft.",
        )

    result = send_campaign_job(db, campaign_id)
    return {
        "message": "Отправка кампании завершена",
        "sent_direct": True,
        "result": result,
    }


@app.get(
    "/campaigns/{campaign_id}/report",
    response_model=schemas.SendSummary,
    summary="Получить отчёт по кампании",
    tags=["Кампании"],
)
def campaign_report(
    campaign_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Возвращает статистику по отправке кампании.
    """
    total = db.scalar(select(func.count()).select_from(models.Message).where(models.Message.campaign_id == campaign_id))
    sent = db.scalar(
        select(func.count()).select_from(models.Message).where(
            models.Message.campaign_id == campaign_id,
            models.Message.status == "sent"
        )
    )
    failed = db.scalar(
        select(func.count()).select_from(models.Message).where(
            models.Message.campaign_id == campaign_id,
            models.Message.status == "failed"
        )
    )
    skipped = db.scalar(
        select(func.count()).select_from(models.Message).where(
            models.Message.campaign_id == campaign_id,
            models.Message.status == "skipped"
        )
    )

    return schemas.SendSummary(
        campaign_id=campaign_id,
        total=total or 0,
        sent=sent or 0,
        failed=failed or 0,
        skipped=skipped or 0,
    )


@app.get(
    "/campaigns/{campaign_id}/report/html",
    response_class=HTMLResponse,
    summary="HTML-отчёт по кампании",
    tags=["Кампании"],
)
def campaign_report_html(
    campaign_id: int,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    HTML-отчёт по кампании с возможностью сохранения в PDF через браузер.
    """
    camp = db.get(models.Campaign, campaign_id)
    if not camp:
        raise HTTPException(status_code=404, detail="Кампания не найдена")

    total = db.scalar(
        select(func.count()).select_from(models.Message).where(models.Message.campaign_id == campaign_id)
    ) or 0

    sent = db.scalar(
        select(func.count()).select_from(models.Message).where(
            models.Message.campaign_id == campaign_id,
            models.Message.status == "sent"
        )
    ) or 0

    failed = db.scalar(
        select(func.count()).select_from(models.Message).where(
            models.Message.campaign_id == campaign_id,
            models.Message.status == "failed"
        )
    ) or 0

    skipped = db.scalar(
        select(func.count()).select_from(models.Message).where(
            models.Message.campaign_id == campaign_id,
            models.Message.status == "skipped"
        )
    ) or 0

    rows = db.execute(
        select(
            models.Message.id,
            models.Contact.email,
            models.Contact.full_name,
            models.Message.status,
            models.Message.attempts,
            models.Message.last_error,
            models.Message.sent_at,
        )
        .join(models.Contact, models.Contact.id == models.Message.contact_id)
        .where(models.Message.campaign_id == campaign_id)
        .order_by(models.Message.id.asc())
    ).all()

    rows_html = ""
    for row in rows:
        sent_at = row.sent_at.strftime("%d.%m.%Y %H:%M:%S") if row.sent_at else "-"
        rows_html += f"""
        <tr>
            <td>{row.id}</td>
            <td>{row.email}</td>
            <td>{row.full_name or "-"}</td>
            <td>{row.status}</td>
            <td>{row.attempts}</td>
            <td>{row.last_error or "-"}</td>
            <td>{sent_at}</td>
        </tr>
        """

    return f"""
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <title>Отчёт по кампании {campaign_id}</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 30px;
                color: #222;
                background: #f8fafc;
            }}
            h1, h2 {{
                color: #0b5cab;
            }}
            .block {{
                margin-bottom: 20px;
                padding: 15px;
                border: 1px solid #dcdcdc;
                border-radius: 10px;
                background: #ffffff;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 10px;
                font-size: 14px;
                background: white;
            }}
            th, td {{
                border: 1px solid #cfcfcf;
                padding: 8px;
                text-align: left;
                vertical-align: top;
            }}
            th {{
                background: #eef4fb;
            }}
            .note {{
                margin-top: 20px;
                padding: 12px;
                background: #fff8e8;
                border: 1px solid #f0d999;
                border-radius: 8px;
            }}
        </style>
    </head>
    <body>
        <h1>Отчёт по кампании массовой рассылки</h1>

        <div class="block">
            <p><strong>ID кампании:</strong> {camp.id}</p>
            <p><strong>Название:</strong> {camp.name}</p>
            <p><strong>Описание:</strong> {camp.description or "-"}</p>
            <p><strong>Статус:</strong> {camp.status}</p>
        </div>

        <div class="block">
            <h2>Сводная информация</h2>
            <p><strong>Всего сообщений:</strong> {total}</p>
            <p><strong>Успешно отправлено:</strong> {sent}</p>
            <p><strong>Ошибок отправки:</strong> {failed}</p>
            <p><strong>Пропущено:</strong> {skipped}</p>
        </div>

        <div class="block">
            <h2>Детализация сообщений</h2>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Email</th>
                        <th>ФИО</th>
                        <th>Статус</th>
                        <th>Попытки</th>
                        <th>Ошибка</th>
                        <th>Время отправки</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </div>

        <div class="note">
            Для сохранения в PDF открой меню браузера и выбери:
            <strong>Печать → Сохранить как PDF</strong>.
        </div>
    </body>
    </html>
    """


# ---------- ОТПИСКА ----------
@app.get("/unsubscribe/{contact_id}", summary="Отписать контакт от рассылки", tags=["Отписка"])
def unsubscribe(contact_id: int, db: Session = Depends(get_db)):
    """
    Отписка контакта от рассылки.
    """
    contact = db.get(models.Contact, contact_id)
    if not contact:
        raise HTTPException(status_code=404, detail="Контакт не найден")

    existing = db.scalar(select(models.Unsubscribe).where(models.Unsubscribe.contact_id == contact_id))
    if existing:
        return {"ok": True, "message": "Контакт уже был отписан ранее"}

    db.add(models.Unsubscribe(contact_id=contact_id, reason="user_request"))
    db.commit()
    return {"ok": True, "message": "Контакт успешно отписан"}